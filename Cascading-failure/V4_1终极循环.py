#1 导包
# import matplotlib.pyplot as plt
import networkx as nx
import numpy as np
import random
import time
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# import network


# p = 0.9


for pp in range(35,100,1):
    p = pp/100
    print('!!!!开始计算{num}!!!!!'.format(num=p))
    perA_list = []
    perB_list = []
    for count_time in range(30):
        # #1 import doc
        print("start calculate N0.", count_time)
        print('程序开始：%s' % time.strftime("%a %b %d %H:%M:%S %Y", time.localtime()))
        graph1 = nx.read_adjlist(r'C:\Users\Ludwigia\Desktop\实验数据\BA-pare\废案\极端差异\graph1.adjlist')
        print('g1读取完毕:%s' % time.strftime("%a %b %d %H:%M:%S %Y", time.localtime()))

        graph2 = nx.read_adjlist(r'C:\Users\Ludwigia\Desktop\实验数据\BA-pare\废案\极端差异\graph2.adjlist')
        print('g2读取完毕：' + 'my name2:%s' % time.strftime("%a %b %d %H:%M:%S %Y", time.localtime()))

        G = nx.read_adjlist(r'C:\Users\Ludwigia\Desktop\实验数据\BA-pare\废案\极端差异\G.adjlist')
        print('G读取完毕：' + 'my name2:%s' % time.strftime("%a %b %d %H:%M:%S %Y", time.localtime()))

        # nodenumberB = 43569561  # 节点数量，就是每个网络的节点总数
        # nodenumberA = 47914
        nodenumberB = graph2.number_of_nodes()
        nodenumberA = graph1.number_of_nodes()
        print(nodenumberA)
        print(nodenumberB)
        run_time = 1

        # print('级联失效开始:%s'%time.strftime("%a %b %d %H:%M:%S %Y", time.localtime()))
        #

        # Part II 级联失效模型
        # 第一步，网络A中 p%的节点受到攻击并被移除
        countp = (1 - p) * nodenumberA
        curA = nodenumberA  # 剩下所有节点数和
        curB = nodenumberB

        while (countp > 0):
            jone = random.randint(0, nodenumberA - 1)  # 编号1-约270000，实际只有47914
            # print(jone)
            if G.has_node('Gone{jone}'.format(jone=jone)):
                G.remove_node('Gone{jone}'.format(jone=jone))
                graph1.remove_node('{jone}'.format(jone=jone))  # 原图也跟着操作，然而命名格式不一样
                countp = countp - 1
                curA = curA - 1
        # jone = 5
        # G.remove_node('Gone{jone}'.format(jone=jone))
        # graph1.remove_node('{jone}'.format(jone=jone)) #原图也跟着操作，然而命名格式不一样
        # jone = 9
        # G.remove_node('Gone{jone}'.format(jone=jone))
        # graph1.remove_node('{jone}'.format(jone=jone))
        # curA = curA - 2

        # 第二步，循环判断A中失效的点
        ##两种方法迭代A网络：一是全遍历，只选取A没失效的点
        ##二是只遍历点集合下标(用这种)
        ###两种方法寻找有无依赖链接：一是G.has_edge; 二是[(u, v) for (u, v, d) in G.edges(data=True) if d['lineweight'] == 2]
        # g = list(G.nodes)
        print("A:{a}".format(a=curA) + " B:{b}".format(b=curB))
        resA = curA  # resA为上一步操作A剩下的节点数，如果与本次操作剩下的节点数curA相同，证明A网络趋于稳定
        resB = curB
        round = 0
        # print('第{num}轮'.format(num=round))
        nodes_to_remove = set()
        Gnodes_to_remove = set()

        while (True):
            round = round + 1
            print('第{num}轮'.format(num=round))
            # print(nx.number_connected_components(G))

            # 第一步，删除A中没有依赖链接的节点
            for node in graph1.nodes():
                flag = 0
                nodeG = 'Gone' + str(node)
                for key in G[nodeG]:
                    if key.startswith('Gtwo'):
                        flag = 1
                        break
                if flag == 0:
                    nodes_to_remove.add(node)
                    Gnodes_to_remove.add(nodeG)
            # 删除没有依赖链接的节点
            G.remove_nodes_from(Gnodes_to_remove)
            graph1.remove_nodes_from(nodes_to_remove)
            nodes_to_remove.clear()
            Gnodes_to_remove.clear()
            curA = len(graph1)  # 更新节点数量

            if curA == 0:
                print(nx.number_connected_components(G))
                print('网络A被摧毁，迭代结束')
                break

            if round > 1 and resA == curA:
                print(nx.number_connected_components(G))
                print("网络趋于平衡,迭代结束")
                break

            print("A:{a}".format(a=curA) + " B:{b}".format(b=curB))
            resA = curA
            resB = curB
            # print(G.edges)

            # 第二步，删除A中没有与GCC相连的节点
            # 创建一个当前网络的节点集合
            current_nodes = set(graph1.nodes())
            # 获取当前网络的最大连通组件
            connected_components = nx.connected_components(graph1)
            gcc = max(connected_components, key=len)
            # 将最大连通组件转换为集合
            gcc_set = set(gcc)
            # 找到需要删除的节点
            nodes_to_remove = current_nodes - gcc_set
            # 删除没有与GCC相连的节点
            graph1.remove_nodes_from(nodes_to_remove)
            node_mapping = {'Gone' + node for node in nodes_to_remove}
            G.remove_nodes_from(node_mapping)
            node_mapping.clear()
            curA = len(graph1)  # 更新节点数量

            if curA == 0:
                print(nx.number_connected_components(G))
                print('网络A被摧毁，迭代结束')
                break

            if round > 1 and resA == curA:
                print(nx.number_connected_components(G))
                print("网络趋于平衡,迭代结束")
                break

            print("A:{a}".format(a=curA) + " B:{b}".format(b=curB))
            resA = curA
            resB = curB
            # print(G.edges)

            # 第三步，删除B中没有与依赖链接的节点
            # nodes_to_remove = []
            # Gnodes_to_remove = []
            for node in graph2.nodes():
                flag = 0
                nodeG = 'Gtwo' + str(node)  # 没错，node里已带有N
                for key in G[nodeG]:
                    if key.startswith('Gone'):
                        flag = 1
                        break
                if flag == 0:
                    nodes_to_remove.add(node)
                    Gnodes_to_remove.add(nodeG)
            # 删除没有依赖链接的节点
            G.remove_nodes_from(Gnodes_to_remove)
            graph2.remove_nodes_from(nodes_to_remove)
            nodes_to_remove.clear()
            Gnodes_to_remove.clear()
            curB = len(graph2)  # 更新节点数量

            if curB == 0:
                print(nx.number_connected_components(G))
                print('网络B被摧毁，迭代结束')
                break

            if round > 1 and resB == curB:
                print(nx.number_connected_components(G))
                print("网络趋于平衡,迭代结束")
                break

            print("A:{a}".format(a=curA) + " B:{b}".format(b=curB))
            resA = curA
            resB = curB
            # print(G.edges)

            # 第四步，删除A中没有与GCC相连的节点
            # 创建一个当前网络的节点集合
            current_nodes = set(graph2.nodes())
            # 获取当前网络的最大连通组件
            connected_components = nx.connected_components(graph2)
            gcc = max(connected_components, key=len)
            # 将最大连通组件转换为集合
            gcc_set = set(gcc)
            # 找到需要删除的节点
            nodes_to_remove = current_nodes - gcc_set
            # 删除没有与GCC相连的节点
            graph2.remove_nodes_from(nodes_to_remove)
            node_mapping = {'Gtwo' + node for node in nodes_to_remove}
            G.remove_nodes_from(node_mapping)
            node_mapping.clear()
            print('B_GCC_generated：%s' % time.strftime("%a %b %d %H:%M:%S %Y", time.localtime()))
            curB = len(graph2)  # 更新节点数量

            if curB == 0:
                print(nx.number_connected_components(G))
                print('网络B被摧毁，迭代结束')
                break

            if round > 1 and resB == curB:
                print(nx.number_connected_components(G))
                print("网络趋于平衡,迭代结束")
                break

            print("A:{a}".format(a=curA) + " B:{b}".format(b=curB))
            resA = curA
            resB = curB
            # print(G.edges)

        print('迭代后检测：')
        # 判断最大gcc占剩余节点比例
        largestA = max(nx.connected_components(graph1), key=len)
        largestB = max(nx.connected_components(graph2), key=len)
        # print(largestB)
        # per = len(largest) / (resA + resB)
        perA = len(largestA) / nodenumberA
        perB = len(largestB) / nodenumberB
        print(len(largestA))
        print(perA)
        print(len(largestB))
        print(perB)
        if perA < 0.1 or perB < 0.1:
            print('网络不存在gcc')
        else:
            print('网络平衡')
        perA_list.append(perA)
        perB_list.append(perB)

    print(perA_list)
    print(perB_list)
    A_MEAN = np.mean(perA_list)
    B_MEAN = np.mean(perB_list)
    try:
        wb = openpyxl.load_workbook('data.xlsx')
        # 如果工作簿中没有名为“Sheet1”的工作表，则创建一个新工作表
        if 'Sheet1' not in wb.sheetnames:
            ws = wb.create_sheet('Sheet1')
        # 如果工作簿中已经有名为“Sheet1”的工作表，则选择该工作表
        else:
            ws = wb['Sheet1']
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
    # 创建数据帧
    df = pd.DataFrame({'p': [p], 'perA': [A_MEAN], 'perB': [B_MEAN]})
    # 将数据帧追加到工作表中
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
    # 保存Excel文件
    wb.save('data.xlsx')

    print('my name2:%s' % time.strftime("%a %b %d %H:%M:%S %Y", time.localtime()))
    # print(G.nodes)
    # print(G.edges)